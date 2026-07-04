"""Image-aware rendering for non-DOCX GMP templates (.xlsx, .vsdx).

The .docx path uses docxtpl, which turns a ``{{ qa_stamp }}`` tag into an
``InlineImage`` natively. Excel and Visio have no such engine, so this module
supplies the equivalent "tag -> real image" behaviour for each format:

  * Excel (.xlsx)  -> render text tags per cell, then ANCHOR the stamp/signature
    PNGs as real pictures at the cell that held the tag (openpyxl). LibreOffice
    Calc renders anchored images to PDF reliably, so this is the production path.

  * Visio (.vsdx)  -> render text tags in the page XML. Raster images embedded
    inside a .vsdx via <ForeignData> are NOT reliably rendered by LibreOffice's
    libvisio importer, so images for Visio are applied as a POST-conversion PDF
    overlay via ``stamp_pdf`` (reportlab + pypdf, the same primitive the
    watermark uses). ``embed_vsdx_foreign_image`` is provided for completeness
    but must be verified against your specific soffice build before relying on it.

All functions take a plain text ``context`` (image keys already blanked, exactly
like ``_build_template_context(template_for_images=None)``) plus, where relevant,
an ``images`` map of ``{tag: absolute_png_path}`` resolved by the caller (e.g.
from ``_resolve_stamp_path`` / ``_resolve_signature_path``). Keeping image
resolution in the caller leaves this module free of Frappe coupling and unit
testable in isolation.
"""

import io
import os
import re
import zipfile
from xml.sax.saxutils import escape as xml_escape

# Matches {{ tag }} with flexible inner whitespace; tag is a bare identifier.
_TOKEN = re.compile(r"{{\s*([A-Za-z0-9_]+)\s*}}")


def _substitute(text, context):
    """Replace every ``{{ tag }}`` with ``str(context[tag])`` (missing -> "")."""
    return _TOKEN.sub(lambda m: str(context.get(m.group(1), "") or ""), text)


def _render_cell_text(text, context):
    """Render a spreadsheet cell string with full Jinja2 so both plain tags
    (``{{ docname }}``) and expressions/conditionals
    (``{{ 'CONTROLLED COPY' if is_active else 'OBSOLETE' }}``) resolve. Jinja2 is
    always present under Frappe; if it is somehow unavailable we degrade to a
    bare-tag token substitution (expressions are then left as-is rather than
    crashing the render). ``autoescape=False`` because spreadsheet cells are not
    HTML — values must land verbatim."""
    if "{{" not in text and "{%" not in text:
        return text
    try:
        from jinja2 import Environment

        return Environment(autoescape=False).from_string(text).render(**context)
    except ImportError:
        return _substitute(text, context)


def _has_tag(text, tag):
    return re.search(r"{{\s*%s\s*}}" % re.escape(tag), text) is not None


def _strip_tag(text, tag):
    return re.sub(r"{{\s*%s\s*}}" % re.escape(tag), "", text)


def _size_image(pic, width_px):
    """Size an openpyxl image to ``width_px`` preserving aspect when the native
    size is known (Pillow present), else fall back to a square box."""
    native_w = getattr(pic, "width", None)
    native_h = getattr(pic, "height", None)
    if native_w and native_h:
        pic.height = int(round(width_px * native_h / native_w))
        pic.width = width_px
    else:
        pic.width = pic.height = width_px


# --------------------------------------------------------------------------- #
#  Excel                                                                       #
# --------------------------------------------------------------------------- #
def render_xlsx(source_path, out_path, context, images=None, stamp_width_px=150):
    """Render an .xlsx template to ``out_path``.

    Text ``{{ tag }}`` cells are rendered in place. Any cell whose text contains
    an image tag (a key of ``images`` with a truthy path) has that token removed
    and the PNG anchored as a floating picture over the cell — so ``{{ qa_stamp }}``
    becomes the actual approved/rejected stamp, and signature tags become the
    actual signature images. Returns ``out_path``.
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    images = {t: p for t, p in (images or {}).items() if p and os.path.exists(p)}
    wb = load_workbook(source_path)
    for ws in wb.worksheets:
        # Collect anchors first; add_image() after the value pass so we don't
        # mutate the picture collection while scanning cells.
        anchors = []
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                raw = cell.value
                hit = next((t for t in images if _has_tag(raw, t)), None)
                if hit:
                    raw = _strip_tag(raw, hit)
                    anchors.append((cell.coordinate, images[hit]))
                rendered = _render_cell_text(raw, context).strip()
                cell.value = rendered or None
        for coord, path in anchors:
            pic = XLImage(path)
            _size_image(pic, stamp_width_px)
            ws.add_image(pic, coord)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
#  Visio                                                                       #
# --------------------------------------------------------------------------- #
def _is_visio_page(name):
    return (
        name.startswith("visio/pages/")
        and name.endswith(".xml")
        and "/_rels/" not in name
        and not name.endswith("pages.xml")
    )


def render_vsdx(source_path, out_path, context):
    """Render .vsdx text tags into ``out_path`` (images handled separately via
    ``stamp_pdf`` after PDF conversion). Substituted values are XML-escaped so a
    field containing ``&``/``<``/``>`` cannot corrupt the package. Non-page parts
    are copied through byte-for-byte. Returns ``out_path``.
    """
    def render_xml(raw):
        return _TOKEN.sub(
            lambda m: xml_escape(str(context.get(m.group(1), "") or "")), raw
        )

    with zipfile.ZipFile(source_path) as zin:
        items = [(info.filename, zin.read(info.filename)) for info in zin.infolist()]
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items:
            if _is_visio_page(name):
                data = render_xml(data.decode("utf-8")).encode("utf-8")
            zout.writestr(name, data)
    return out_path


# --------------------------------------------------------------------------- #
#  PDF image overlay (reliable image path for Visio; generic for any PDF)      #
# --------------------------------------------------------------------------- #
def stamp_pdf(pdf_bytes, placements):
    """Overlay PNGs onto an existing PDF and return the new PDF bytes.

    ``placements`` is a list of dicts::

        {"page": 0, "image": "/abs/qaapproved.png",
         "x": 400, "y": 60, "width": 120}

    Coordinates are PDF points with a bottom-left origin (reportlab convention);
    ``width`` is in points and height is derived from the image aspect ratio.
    ``mask='auto'`` honours PNG transparency. This mirrors ``_apply_watermark``
    in the controller and is the recommended way to place the QA stamp on
    Visio-sourced PDFs, where in-package image embedding is unreliable.
    """
    from pypdf import PdfReader, PdfWriter
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas

    by_page = {}
    for pl in placements:
        by_page.setdefault(pl.get("page", 0), []).append(pl)

    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    for idx, page in enumerate(reader.pages):
        for pl in by_page.get(idx, []):
            if not pl.get("image") or not os.path.exists(pl["image"]):
                continue
            width = float(page.mediabox.width)
            height = float(page.mediabox.height)
            buf = io.BytesIO()
            c = canvas.Canvas(buf, pagesize=(width, height))
            img = ImageReader(pl["image"])
            iw, ih = img.getSize()
            w = float(pl.get("width", 120))
            h = w * ih / iw
            c.drawImage(img, float(pl["x"]), float(pl["y"]), width=w, height=h, mask="auto")
            c.save()
            buf.seek(0)
            page.merge_page(PdfReader(buf).pages[0])
        writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
#  Native Visio image embed — OPTIONAL, verify against your soffice build      #
# --------------------------------------------------------------------------- #
def embed_vsdx_foreign_image(source_path, out_path, image_path, page_rels="visio/pages/_rels/page1.xml.rels"):
    """Inject ``image_path`` into a .vsdx as a media part wired to a
    <ForeignData> shape reference.

    NOTE: LibreOffice's libvisio importer does not reliably rasterise
    ForeignData bitmaps when converting to PDF. Treat this as experimental and
    verify the produced PDF on your server before using it in the GMP flow;
    prefer ``stamp_pdf`` for a guaranteed result. Provided so the fully-native
    approach is documented in code rather than only in prose.
    """
    media_name = "visio/media/qa_stamp.png"
    with open(image_path, "rb") as fh:
        media_bytes = fh.read()

    with zipfile.ZipFile(source_path) as zin:
        items = {info.filename: zin.read(info.filename) for info in zin.infolist()}

    # 1. register png content type
    ct = items["[Content_Types].xml"].decode("utf-8")
    if 'Extension="png"' not in ct:
        ct = ct.replace(
            "</Types>",
            '  <Default Extension="png" ContentType="image/png"/>\n</Types>',
        )
    items["[Content_Types].xml"] = ct.encode("utf-8")

    # 2. relationship from the page to the media part
    rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    img_rel_type = "http://schemas.microsoft.com/visio/2010/relationships/image"
    if page_rels in items:
        rels = items[page_rels].decode("utf-8")
    else:
        rels = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="{rels_ns}"></Relationships>'
    rels = rels.replace(
        "</Relationships>",
        f'  <Relationship Id="rIdStamp" Type="{img_rel_type}" Target="../media/qa_stamp.png"/>\n</Relationships>',
    )
    items[page_rels] = rels.encode("utf-8")

    # 3. the media bytes
    items[media_name] = media_bytes

    # 4. a ForeignData shape referencing the relationship
    page = "visio/pages/page1.xml"
    xml = items[page].decode("utf-8")
    shape = (
        '    <Shape ID="900" Type="Foreign">\n'
        '      <Cell N="PinX" V="4.25"/><Cell N="PinY" V="2.2"/>\n'
        '      <Cell N="Width" V="1.5"/><Cell N="Height" V="1.5"/>\n'
        '      <Cell N="LocPinX" V="0.75"/><Cell N="LocPinY" V="0.75"/>\n'
        '      <ForeignData ForeignType="Bitmap"><Rel r:id="rIdStamp"/></ForeignData>\n'
        "    </Shape>\n"
    )
    xml = xml.replace("</Shapes>", shape + "  </Shapes>")
    items[page] = xml.encode("utf-8")

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)
    return out_path
